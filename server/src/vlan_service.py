"""Plan d'adressage fourni par l'équipe réseau : plage d'adresses → VLAN.

Pourquoi une table de **plages** et non une liste d'hôtes : une machine sur
port d'accès ne peut pas connaître son VLAN, mais l'agent remonte son adresse
IP à chaque battement. Un VLAN correspondant presque toujours à une ou
plusieurs plages contiguës, la table permet de déduire le VLAN de **tout** le
parc sans saisie par hôte — et la déduction suit d'elle-même quand une machine
change d'adresse ou de site.

Une liste `hôte → VLAN` serait juste le jour de son export et fausse dès la
première machine rebranchée.

**La plage est la forme canonique, pas le CIDR.** Les équipes réseau
raisonnent en « de telle adresse à telle adresse » : `10.20.4.1 - 10.20.4.254`
ne s'écrit pas en CIDR sans le déformer (ce n'est pas `10.20.4.0/24`, qui
inclut l'adresse réseau et la diffusion). Tout est donc ramené à un couple
première/dernière adresse, dont le CIDR n'est qu'un cas particulier.

Le fichier vient d'Excel : séparateur point-virgule, signature d'octets,
en-têtes accentués, plage tantôt dans une colonne tantôt dans deux. Tout cela
est absorbé ici plutôt que d'être exigé de l'équipe réseau.
"""

from __future__ import annotations

import csv
import io
import ipaddress
import re
from dataclasses import dataclass
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

#: Bornes 802.1Q. 0 et 4095 sont réservés.
VLAN_MIN = 1
VLAN_MAX = 4094

#: En-têtes acceptés, en minuscules et sans accent.
_SPAN_HEADERS = {
    "sous-reseau", "sous reseau", "subnet", "reseau", "cidr", "network",
    "plage", "plage ip", "plage d'adresses", "adresses", "range", "ip range",
}
_VLAN_HEADERS = {"vlan", "vlan id", "vlanid", "id vlan", "numero", "numero vlan", "tag"}
_LABEL_HEADERS = {"libelle", "label", "nom", "name", "description", "designation", "zone"}

#: Certaines équipes livrent la plage en deux colonnes plutôt qu'une.
_START_HEADERS = {"debut", "ip debut", "adresse debut", "start", "start ip", "ip start", "premiere ip", "de"}
_END_HEADERS = {"fin", "ip fin", "adresse fin", "end", "end ip", "ip end", "derniere ip"}

#: Séparateurs de plage rencontrés dans les exports.
_RANGE_SPLIT = re.compile(r"\s*(?:-|–|—|\bto\b|\ba\b)\s*", re.IGNORECASE)

_ACCENTS = str.maketrans("àâäéèêëîïôöùûüç", "aaaeeeeiioouuuc")

#: Espaces exotiques produits par Excel (insécable, fine).
_ODD_SPACES = str.maketrans({" ": " ", " ": " ", " ": " "})


class VlanImportError(ValueError):
    """Fichier inexploitable — message destiné à l'exploitant."""


@dataclass(frozen=True)
class SubnetRow:
    """Une plage du plan d'adressage.

    `cidr` porte la forme lisible telle qu'elle sera affichée et sert
    d'identité ; `range_start` / `range_end` portent la forme sur laquelle on
    calcule. Les deux sont conservés : afficher `10.20.4.1-10.20.4.254` sous
    la forme `10.20.4.0/24` mentirait sur ce que l'équipe réseau a écrit.
    """

    cidr: str
    vlan: str
    label: Optional[str] = None
    range_start: Optional[str] = None
    range_end: Optional[str] = None

    @property
    def span(self) -> Optional[Tuple[int, int]]:
        if self.range_start and self.range_end:
            try:
                return (
                    int(ipaddress.ip_address(self.range_start)),
                    int(ipaddress.ip_address(self.range_end)),
                )
            except ValueError:
                return None
        # Ligne d'une version antérieure, sans bornes stockées.
        parsed = parse_span(self.cidr)
        return (parsed[0], parsed[1]) if parsed else None

    @property
    def size(self) -> int:
        bounds = self.span
        return (bounds[1] - bounds[0] + 1) if bounds else 0


@dataclass
class ImportReport:
    rows: List[SubnetRow]
    rejected: List[Dict[str, Any]]

    @property
    def accepted_count(self) -> int:
        return len(self.rows)


def _fold(value: Any) -> str:
    return str(value or "").strip().lower().translate(_ACCENTS)


def _clean(value: Any) -> str:
    return str(value or "").translate(_ODD_SPACES).strip()


def parse_span(value: Any) -> Optional[Tuple[int, int, str]]:
    """Ramène une saisie réseau à (première, dernière, forme lisible).

    Accepte, dans l'ordre d'essai :
      * `10.20.4.1 - 10.20.4.254` — la forme que les équipes réseau écrivent ;
      * `10.20.4.0/24` — notation CIDR ;
      * `10.20.4.0 255.255.255.0` — réseau et masque ;
      * `10.20.4.5` — adresse seule.
    """
    raw = _clean(value)
    if not raw:
        return None

    # Plage explicite. Testée d'abord : « 10.20.4.1 - 10.20.4.254 » contient
    # des séparateurs qui égareraient les autres formes.
    parts = [p for p in _RANGE_SPLIT.split(raw) if p]
    if len(parts) == 2:
        try:
            first = ipaddress.ip_address(parts[0].strip())
            last = ipaddress.ip_address(parts[1].strip())
        except ValueError:
            first = last = None
        if first is not None and last is not None and first.version == last.version:
            if int(first) > int(last):
                first, last = last, first
            return int(first), int(last), "%s-%s" % (first, last)

    # Réseau et masque, en deux mots.
    words = raw.split()
    if len(words) == 2:
        try:
            network = ipaddress.ip_network("%s/%s" % (words[0], words[1]), strict=False)
            return (
                int(network.network_address),
                int(network.broadcast_address),
                str(network),
            )
        except ValueError:
            return None

    # CIDR, ou adresse seule.
    try:
        network = ipaddress.ip_network(raw, strict=False)
    except ValueError:
        return None
    return int(network.network_address), int(network.broadcast_address), str(network)


def normalise_cidr(value: Any) -> Optional[str]:
    """Forme lisible d'une saisie réseau, ou None. Conservée pour l'appelant."""
    parsed = parse_span(value)
    return parsed[2] if parsed else None


def normalise_vlan(value: Any) -> Optional[str]:
    """Extrait un identifiant 802.1Q valide, ou None.

    Tolère « VLAN 20 » et « 20 (Monétique) » : l'équipe réseau écrit rarement
    un entier nu dans un tableur.
    """
    raw = _clean(value)
    if not raw:
        return None
    match = re.search(r"\d+", raw)
    if not match:
        return None
    number = int(match.group())
    if not (VLAN_MIN <= number <= VLAN_MAX):
        return None
    return str(number)


def _read_csv(content: bytes) -> List[Sequence[Any]]:
    text = content.decode("utf-8-sig", errors="replace")
    if not text.strip():
        return []
    sample = text[:4096]
    try:
        # Excel français écrit des points-virgules ; l'anglais des virgules.
        delimiter = csv.Sniffer().sniff(sample, delimiters=";,\t").delimiter
    except csv.Error:
        delimiter = ";" if sample.count(";") >= sample.count(",") else ","
    return [
        row
        for row in csv.reader(io.StringIO(text), delimiter=delimiter)
        if any(str(c).strip() for c in row)
    ]


def _read_xlsx(content: bytes) -> List[Sequence[Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover - dépendance déclarée
        raise VlanImportError(
            "Lecture .xlsx indisponible sur ce serveur. Réenregistrer le "
            "fichier au format CSV et le réimporter."
        )
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise VlanImportError("Classeur illisible : %s" % exc)
    sheet = workbook.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        if row and any(str(c).strip() for c in row if c is not None):
            rows.append(list(row))
    workbook.close()
    return rows


@dataclass(frozen=True)
class _Columns:
    """Où lire chaque valeur. `span` OU (`start`, `end`) est renseigné."""

    vlan: int
    span: Optional[int] = None
    start: Optional[int] = None
    end: Optional[int] = None
    label: Optional[int] = None


def _column_positions(header: Sequence[Any]) -> Optional[_Columns]:
    """Repère les colonnes d'après leur intitulé, si la ligne en est un."""
    folded = [_fold(c) for c in header]
    span = vlan = label = start = end = None
    for index, name in enumerate(folded):
        if span is None and name in _SPAN_HEADERS:
            span = index
        elif vlan is None and name in _VLAN_HEADERS:
            vlan = index
        elif start is None and name in _START_HEADERS:
            start = index
        elif end is None and name in _END_HEADERS:
            end = index
        elif label is None and name in _LABEL_HEADERS:
            label = index

    if vlan is None:
        return None
    if span is not None:
        return _Columns(vlan=vlan, span=span, label=label)
    if start is not None and end is not None:
        return _Columns(vlan=vlan, start=start, end=end, label=label)
    return None


def parse(content: bytes, filename: str = "") -> ImportReport:
    """Lit le fichier de l'équipe réseau et rend les lignes exploitables.

    Une ligne fautive est **rejetée nommément**, jamais ignorée en silence :
    un import qui avale à moitié un plan de segmentation laisserait des hôtes
    rattachés à un VLAN qui n'est pas le leur, sans que personne ne le sache.
    """
    if not content:
        raise VlanImportError("Fichier vide.")

    lowered = (filename or "").lower()
    if lowered.endswith((".xlsx", ".xlsm")):
        raw_rows = _read_xlsx(content)
    elif lowered.endswith(".xls"):
        raise VlanImportError(
            "Format .xls (Excel 97) non pris en charge. Réenregistrer en .xlsx "
            "ou en CSV."
        )
    else:
        raw_rows = _read_csv(content)

    if not raw_rows:
        raise VlanImportError("Aucune ligne exploitable dans le fichier.")

    columns = _column_positions(raw_rows[0])
    if columns is not None:
        body = raw_rows[1:]
        first_line = 2
    else:
        # Pas d'en-tête reconnu : on suppose l'ordre du document demandé —
        # plage, VLAN, libellé.
        columns = _Columns(vlan=1, span=0, label=2)
        body = raw_rows
        first_line = 1

    accepted: List[SubnetRow] = []
    rejected: List[Dict[str, Any]] = []
    seen: Dict[str, int] = {}

    for offset, row in enumerate(body):
        line = offset + first_line

        def cell(index: Optional[int]) -> Any:
            if index is None or index >= len(row):
                return None
            return row[index]

        if columns.span is not None:
            raw_span = _clean(cell(columns.span))
            parsed = parse_span(raw_span)
        else:
            raw_span = "%s-%s" % (_clean(cell(columns.start)), _clean(cell(columns.end)))
            parsed = parse_span(raw_span)

        vlan = normalise_vlan(cell(columns.vlan))

        if parsed is None:
            rejected.append(
                {"line": line, "reason": "plage d'adresses illisible", "value": raw_span}
            )
            continue
        if vlan is None:
            rejected.append(
                {
                    "line": line,
                    "reason": "VLAN absent ou hors 1-4094",
                    "value": _clean(cell(columns.vlan)),
                }
            )
            continue

        start_int, end_int, display = parsed
        if display in seen:
            rejected.append(
                {
                    "line": line,
                    "reason": "plage déjà définie ligne %d" % seen[display],
                    "value": display,
                }
            )
            continue

        seen[display] = line
        label = _clean(cell(columns.label)) or None
        accepted.append(
            SubnetRow(
                cidr=display,
                vlan=vlan,
                label=label[:120] if label else None,
                range_start=str(ipaddress.ip_address(start_int)),
                range_end=str(ipaddress.ip_address(end_int)),
            )
        )

    if not accepted:
        raise VlanImportError(
            "Aucune ligne valide. Attendu : une plage (10.20.4.1-10.20.4.254 "
            "ou 10.20.4.0/24), un VLAN (1-4094), un libellé facultatif."
        )
    return ImportReport(rows=accepted, rejected=rejected)


def match_ip(ip_address: Optional[str], rows: Iterable[SubnetRow]) -> Optional[SubnetRow]:
    """Plage la plus étroite contenant cette adresse.

    La plus étroite, et non la première trouvée : les plans d'adressage
    déclarent couramment la plage d'un site *et* les sous-plages qui la
    découpent. Rendre la plage large rattacherait tous les hôtes au mauvais
    VLAN.
    """
    if not ip_address:
        return None
    try:
        address = int(ipaddress.ip_address(_clean(ip_address)))
    except ValueError:
        return None

    best: Optional[SubnetRow] = None
    best_size = -1
    for row in rows:
        bounds = row.span
        if bounds is None:
            continue
        start, end = bounds
        if not (start <= address <= end):
            continue
        size = end - start + 1
        if best_size < 0 or size < best_size:
            best = row
            best_size = size
    return best
